"""
独立脚本：爬取 3月9日-3月15日 全地区游戏榜单并导出 Excel，同时 AI 分析异动原因。
用法：cd D:/project/content-monitor && python fetch_weekly_rank.py
"""

import sys
import os
import time
import logging
from datetime import date, timedelta

# 确保能 import 项目模块
sys.path.insert(0, os.path.dirname(__file__))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


# ── 配置 ──────────────────────────────────────────────────────────────
START_DATE = date(2026, 3, 9)
END_DATE   = date(2026, 3, 15)

ALL_COUNTRIES = list(__import__("scrapers.qimai_rank", fromlist=["REGIONS"]).REGIONS.keys())
# ALL_COUNTRIES = ["cn", "hk", "tw", "us", "jp", "kr"]  # 若只需要部分地区，取消注释此行

CHART_TYPES = ["free", "grossing"]
TOP = 10

OUTPUT_FILE = f"qimai_rank_{START_DATE.strftime('%m%d')}-{END_DATE.strftime('%m%d')}.xlsx"
# ─────────────────────────────────────────────────────────────────────


def fetch_rank_range():
    """按指定日期范围爬取所有地区榜单，生成自定义 Excel。"""
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from scrapers.qimai_rank import (
        fetch_rank_by_date, REGIONS, CHART_TYPES as CT_MAP,
    )

    # 生成日期列表
    days = []
    d = START_DATE
    while d <= END_DATE:
        days.append(d.strftime("%Y-%m-%d"))
        d += timedelta(days=1)

    date_range_label = f"{START_DATE.strftime('%m月%d日')}-{END_DATE.strftime('%m月%d日')}"

    # 配色
    C_TITLE_BG = "0F1824"
    C_TITLE_FG = "F5B820"
    C_HEAD_BG  = "1A2840"
    C_HEAD_FG  = "F5B820"
    C_RANK_BG  = "0F1824"
    C_RANK_FG  = "6A8090"
    C_ODD_BG   = "141E2E"
    C_EVEN_BG  = "0F1824"
    C_RANK1_FG = "FFD700"
    C_RANK2_FG = "C0C0C0"
    C_RANK3_FG = "CD7F32"
    C_TEXT     = "C8D4E0"
    C_NEW_FG   = "A78BFA"
    C_NEW_BG   = "1E1535"
    C_BORDER   = "243450"

    thin   = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style(c, bg, fg, bold=False, size=10, h_align="center", wrap=False):
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(color=fg, bold=bold, size=size)
        c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
        c.border = border

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_anomalies = []

    total = len(ALL_COUNTRIES) * len(CHART_TYPES)
    done  = 0

    for country in ALL_COUNTRIES:
        cname = REGIONS.get(country, {}).get("name", country)
        for chart_type in CHART_TYPES:
            done += 1
            chart_name = CT_MAP.get(chart_type, chart_type)
            logger.info("[%d/%d] %s %s", done, total, cname, chart_name)

            # 抓取每天数据
            daily_data = []  # [{rank_int: game_name}]
            for d in days:
                raw = fetch_rank_by_date(country, chart_type, d, TOP)
                rank_to_name = {int(v): k for k, v in raw.items() if v is not None}
                daily_data.append(rank_to_name)
                time.sleep(0.6)   # 礼貌爬取

            # 创建 sheet
            sheet_name = f"{cname}_{chart_name}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_view.showGridLines = False

            n_cols = 1 + len(days)

            # 标题行
            title = f"{cname} App Store 游戏{chart_name} TOP{TOP}（{date_range_label}）"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
            tc = ws.cell(row=1, column=1, value=title)
            style(tc, C_TITLE_BG, C_TITLE_FG, bold=True, size=11)
            ws.row_dimensions[1].height = 30

            # 列头行
            rc = ws.cell(row=2, column=1, value="排名")
            style(rc, C_HEAD_BG, C_HEAD_FG, bold=True)
            for ci, d in enumerate(days, 2):
                m, day = d[5:7], d[8:]
                label = f"{int(m)}月{int(day)}日"
                hc = ws.cell(row=2, column=ci, value=label)
                style(hc, C_HEAD_BG, C_HEAD_FG, bold=True)
            ws.row_dimensions[2].height = 26

            # 数据行
            for rank in range(1, TOP + 1):
                row    = rank + 2
                is_odd = rank % 2 == 1
                row_bg = C_ODD_BG if is_odd else C_EVEN_BG

                rank_fg = (C_RANK1_FG if rank == 1 else
                           C_RANK2_FG if rank == 2 else
                           C_RANK3_FG if rank == 3 else C_RANK_FG)
                nc = ws.cell(row=row, column=1, value=rank)
                style(nc, C_RANK_BG, rank_fg, bold=(rank <= 3), size=11)
                ws.row_dimensions[row].height = 22

                for di, day_dict in enumerate(daily_data):
                    game    = day_dict.get(rank, "")
                    is_last = (di == len(daily_data) - 1)
                    is_new  = is_last and game and (game not in daily_data[0].values())

                    gc = ws.cell(row=row, column=di + 2, value=game)
                    if is_new:
                        style(gc, C_NEW_BG, C_NEW_FG, bold=True)
                    else:
                        cell_fg = (C_RANK1_FG if rank == 1 else
                                   C_RANK2_FG if rank == 2 else
                                   C_RANK3_FG if rank == 3 else C_TEXT)
                        style(gc, row_bg, cell_fg,
                              bold=(rank <= 3 and bool(game)))

                    # 异动检测（最后一天对比第一天）
                    if is_last and game:
                        if is_new:
                            all_anomalies.append({
                                "app": game, "region": cname, "chart": chart_name,
                                "change": "★新上榜", "rank": rank, "country": country,
                            })
                        else:
                            first_rank = next(
                                (r for r, n in daily_data[0].items() if n == game), None
                            )
                            if first_rank and first_rank - rank >= 3:
                                all_anomalies.append({
                                    "app": game, "region": cname, "chart": chart_name,
                                    "change": f"↑{first_rank - rank}",
                                    "rank": rank, "country": country,
                                })

            # 底部来源行
            src_row = TOP + 3
            ws.merge_cells(start_row=src_row, start_column=1,
                           end_row=src_row, end_column=n_cols)
            sc = ws.cell(row=src_row, column=1, value="资料来源：七麦数据")
            style(sc, C_TITLE_BG, "445566", size=9, h_align="left")
            ws.row_dimensions[src_row].height = 18

            # 列宽
            ws.column_dimensions["A"].width = 6
            for ci in range(2, n_cols + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 16

            ws.freeze_panes = "B3"

    # 按变化幅度排序异动
    def _sort_key(x):
        c = x.get("change", "")
        if c.startswith("↑"):
            try:
                return int(c[1:])
            except ValueError:
                pass
        return 0

    all_anomalies.sort(key=_sort_key, reverse=True)

    wb.save(OUTPUT_FILE)
    logger.info("Excel 已保存: %s", OUTPUT_FILE)
    return all_anomalies


def analyze_anomalies(anomalies):
    """用 AI 批量分析异动原因，打印结果。"""
    if not anomalies:
        print("\n未检测到明显榜单异动。")
        return

    print(f"\n共检测到 {len(anomalies)} 条榜单异动，开始 AI 分析（最多分析前10条）...\n")

    from qwen_client import analyze_rank_anomaly

    top_anomalies = anomalies[:10]
    results = []

    for i, a in enumerate(top_anomalies, 1):
        app    = a["app"]
        region = a["region"]
        chart  = a["chart"]
        change = a["change"]
        rank   = a["rank"]

        try:
            print(f"[{i}/{len(top_anomalies)}] 分析 《{app}》 {region} {chart} {change} → 第{rank}名 ...")
        except UnicodeEncodeError:
            print(f"[{i}/{len(top_anomalies)}] 分析中...")
        reason = analyze_rank_anomaly(
            app_name=app,
            region=region,
            chart=chart,
            change=change,
            rank=rank,
        )
        results.append({"app": app, "region": region, "chart": chart,
                        "change": change, "rank": rank, "reason": reason})
        time.sleep(1)

    # 打印汇总
    print("\n" + "=" * 70)
    print("榜单异动 AI 分析结果")
    print("=" * 70)
    for r in results:
        try:
            print(f"\n【{r['app']}】{r['region']} {r['chart']}  {r['change']} → 第{r['rank']}名")
            print(f"  {r['reason']}")
        except UnicodeEncodeError:
            pass
    print("=" * 70)

    return results


if __name__ == "__main__":
    print(f"爬取范围：{START_DATE} ~ {END_DATE}，地区：{len(ALL_COUNTRIES)} 个，榜单：{CHART_TYPES}")
    print(f"预计需要约 {len(ALL_COUNTRIES) * len(CHART_TYPES) * len(list(range(int((END_DATE - START_DATE).days + 1)))) * 0.6:.0f} 秒...\n")

    anomalies = fetch_rank_range()
    analyze_anomalies(anomalies)

    print(f"\n完成！Excel 文件：{os.path.abspath(OUTPUT_FILE)}")
