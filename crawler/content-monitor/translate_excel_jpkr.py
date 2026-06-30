"""
读取已生成的 Excel，将日本/韩国 Sheet 中的游戏名翻译为中文，原地更新保存。
用法：cd D:/project/content-monitor && python translate_excel_jpkr.py
"""

import sys
import os
import re
import json
import logging
import time

sys.path.insert(0, os.path.dirname(__file__))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

EXCEL_FILE = "D:/project/content-monitor/qimai_rank_0309-0315.xlsx"

# 需要翻译的 Sheet 名关键词（日本、韩国）
JP_KR_KEYWORDS = ["日本", "韩国"]

def needs_translate(name: str) -> bool:
    """判断是否包含日韩文字。"""
    return bool(re.search(r'[\u3040-\u30FF\uAC00-\uD7AF\u3131-\u318E\u4E00-\u9FFF]', name))

def is_pure_latin(name: str) -> bool:
    """纯英文/数字/符号，不需翻译。"""
    return bool(re.match(r'^[\x00-\x7F\s\!\?\.\,\:\;\-\_\(\)\[\]\&\#\%]+$', name))

def translate_batch(names: list) -> dict:
    """调用 qwen-max 批量翻译游戏名，返回 {原名: 中文名}。"""
    from qwen_client import get_qwen_client
    client = get_qwen_client()

    trans_map = {}
    batch_size = 40

    for i in range(0, len(names), batch_size):
        batch = names[i:i + batch_size]
        prompt = (
            "请将以下游戏名称翻译为中文（优先使用官方中文名，无官方名则音译或意译）。"
            "直接返回 JSON 对象，格式：{\"原名\": \"中文名\", ...}，不要其他内容，不要 markdown。\n\n"
            + "\n".join(f"- {n}" for n in batch)
        )
        logger.info("翻译第 %d 批（%d 个名称）...", i // batch_size + 1, len(batch))
        try:
            completion = client.chat.completions.create(
                model="qwen-max",
                messages=[{"role": "user", "content": prompt}],
                extra_body={"enable_thinking": False},
                max_tokens=2000,
                timeout=60,
            )
            raw = (completion.choices[0].message.content or "").strip()
            # 去除可能的 markdown 代码块
            fence = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', raw)
            if fence:
                raw = fence.group(1).strip()
            batch_map = json.loads(raw)
            trans_map.update(batch_map)
            logger.info("  → 成功翻译 %d 个", len(batch_map))
        except Exception as e:
            logger.warning("  翻译批次失败: %s", e)
        time.sleep(1)

    return trans_map


def main():
    import openpyxl

    logger.info("加载 Excel: %s", EXCEL_FILE)
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # 找出日韩 Sheet
    target_sheets = [
        ws for ws in wb.worksheets
        if any(kw in ws.title for kw in JP_KR_KEYWORDS)
    ]
    logger.info("找到 %d 个日韩 Sheet: %s", len(target_sheets),
                [ws.title for ws in target_sheets])

    # 收集所有需要翻译的游戏名（去重）
    to_translate = set()
    for ws in target_sheets:
        for row in ws.iter_rows(min_row=3):  # 跳过标题行和列头行
            for cell in row:
                val = str(cell.value or "").strip()
                if val and val not in ("-", "资料来源：七麦数据") and not val.isdigit():
                    # 排除纯 ASCII（英文游戏名不需要翻译）
                    if not is_pure_latin(val):
                        to_translate.add(val)

    if not to_translate:
        logger.info("没有需要翻译的内容，退出。")
        return

    names_list = sorted(to_translate)
    logger.info("共 %d 个不重复游戏名需要翻译", len(names_list))

    # AI 翻译
    trans_map = translate_batch(names_list)
    logger.info("翻译完成，共获得 %d 条映射", len(trans_map))

    # 回写到 Excel
    updated = 0
    for ws in target_sheets:
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                val = str(cell.value or "").strip()
                if val in trans_map and trans_map[val] != val:
                    cell.value = trans_map[val]
                    updated += 1

    wb.save(EXCEL_FILE)
    logger.info("已更新 %d 个单元格，保存到: %s", updated, EXCEL_FILE)

    # 打印翻译对照（方便核查）
    print("\n--- 翻译对照表 ---")
    for orig, zh in sorted(trans_map.items()):
        if orig != zh:
            line = f"  {orig}  ->  {zh}"
            try:
                print(line)
            except UnicodeEncodeError:
                print(line.encode("utf-8", errors="replace").decode("utf-8", errors="replace"))


if __name__ == "__main__":
    main()
