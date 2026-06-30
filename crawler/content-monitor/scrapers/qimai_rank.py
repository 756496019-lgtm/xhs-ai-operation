"""七麦数据 (qimai.cn) 游戏榜单爬虫。

爬取七麦开放 API，获取各地区 App Store 游戏免费榜/畅销榜 TOP10，
生成类似周报 PDF 中的表格数据。
"""

import json
import logging
import os
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional

import requests

logger = logging.getLogger(__name__)

_API_BASE = "https://api.qimai.cn/rank/index"

# 缓存文件路径（同目录下）
_CACHE_FILE = os.path.join(os.path.dirname(__file__), ".qimai_cache.json")
# 缓存有效期：6小时（秒）
_CACHE_TTL = 6 * 3600

_API_BASE = "https://api.qimai.cn/rank/index"

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.qimai.cn/",
    "Cookie": (
        "qm_check=A1sdRUIQChtxen8pI0dAMRcOUFseEHBeQF0JTjVBWDgIVh9mfRAQZEoKRiAaGQkSXFRBEBEATAhVVldfPExPHBd2WUtVV0xWIlZZWA8JagJtABlAR2dQOVdEWD1JcQYDGRscFlsNAwQDWkNYBRsCHAkcBBRVRBo%3D; "
        "PHPSESSID=5be4jvic5bi221ar6lcdk7cnb6; "
        "gr_user_id=c84caf3d-9846-4f2e-be8e-547a56d0155e; "
        "ada35577182650f1_gr_session_id=ded42ff1-2f50-421b-8646-ab654a41bcbc; "
        "ada35577182650f1_gr_session_id_sent_vst=ded42ff1-2f50-421b-8646-ab654a41bcbc; "
        "USERINFO=RrS%2BfEQ1ijCHAV55XHZfN%2FMUn2dFNZU%2BicQFas0TpyTxDDlpN%2BScxIFPKZ9ooNEsTGOPBB4NTLt1crFniSlAMdzOQO0%2BnplvW0Iyb50SHnUw%2F201XxjWOSGpSlGVYnVtf5FX5R0tjw5vC1bLhSl8Dw%3D%3D; "
        "AUTHKEY=CxxVpjQzCRlDhA0dru%2BQzHuf0c9LEV8W%2FNe5eSAM6OUI3rVbbumOQ5o2q8tKYkAtb4kdY8Qz0HcnlbeOTxYassgbIxFd9DAUFj8fG8xDyzBW%2Byeu%2FNahEg%3D%3D; "
        "ada35577182650f1_gr_last_sent_sid_with_cs1=ded42ff1-2f50-421b-8646-ab654a41bcbc; "
        "ada35577182650f1_gr_last_sent_cs1=qm20425157953; "
        "ada35577182650f1_gr_cs1=qm20425157953; "
        "aso_ucenter=b5dfUkeQvi9XmpE13Ve9qaDVh3uzkES%2FPbplFGoXIN1H6z88jViXc1f2Gp%2FfDU9ikbA; "
        "synct=1773049528.878; syncd=-219"
    ),
}

# 支持的地区配置
REGIONS = {
    "cn":  {"name": "中国大陆", "name_en": "China"},
    "hk":  {"name": "中国香港", "name_en": "Hong Kong"},
    "tw":  {"name": "中国台湾", "name_en": "Taiwan"},
    "us":  {"name": "美国",     "name_en": "United States"},
    "jp":  {"name": "日本",     "name_en": "Japan"},
    "kr":  {"name": "韩国",     "name_en": "South Korea"},
    "gb":  {"name": "英国",     "name_en": "United Kingdom"},
    "de":  {"name": "德国",     "name_en": "Germany"},
    "fr":  {"name": "法国",     "name_en": "France"},
    "sg":  {"name": "新加坡",   "name_en": "Singapore"},
    "th":  {"name": "泰国",     "name_en": "Thailand"},
    "sa":  {"name": "沙特",     "name_en": "Saudi Arabia"},
    "tr":  {"name": "土耳其",   "name_en": "Turkey"},
    "br":  {"name": "巴西",     "name_en": "Brazil"},
}

# 榜单类型
CHART_TYPES = {
    "free":     "免费榜",
    "grossing": "畅销榜",
    "paid":     "付费榜",
}

# 游戏类目 ID（iOS App Store）
GENRE_GAMES = 6014   # 所有游戏


def _load_cache() -> dict:
    try:
        if os.path.exists(_CACHE_FILE):
            with open(_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def _save_cache(data: dict):
    try:
        with open(_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        logger.warning("缓存写入失败: %s", e)


def _get_rank_change_str(rank_b: Any) -> str:
    """
    将七麦 rank_b 字段转为变化字符串。
    rank_b: {'ranking': 1, 'change': 0, 'genre': '游戏'}
    """
    if not isinstance(rank_b, dict):
        return "-"
    change = rank_b.get("change", 0)
    if change > 0:
        return f"↑{change}"
    elif change < 0:
        return f"↓{abs(change)}"
    else:
        return "-"


def _get_last_rank(rank_b: Any) -> str:
    """获取上周排名。"""
    if not isinstance(rank_b, dict):
        return "-"
    ranking = rank_b.get("ranking")
    if ranking:
        return str(ranking)
    return "-"


def fetch_rank(
    country: str = "cn",
    chart_type: str = "free",
    top: int = 10,
    device: str = "iphone",
) -> Dict[str, Any]:
    """
    爬取指定地区和类型的游戏榜单。

    Returns:
        {
            "country": "cn",
            "country_name": "中国大陆",
            "chart_type": "free",
            "chart_name": "免费榜",
            "fetch_time": "2026-03-09T...",
            "items": [
                {
                    "rank": 1,
                    "app_name": "王者荣耀",
                    "app_id": "...",
                    "publisher": "...",
                    "icon": "...",
                    "last_rank": "4",
                    "change": "↑3",
                    "genre": "动作",
                    "is_chinese": False,  # 是否中国厂商
                },
                ...
            ]
        }
    """
    params = {
        "country": country,
        "genre":   GENRE_GAMES,
        "device":  device,
        "brand":   chart_type,   # 正确参数：brand=free/grossing（type 参数无效）
    }
    items = []
    try:
        resp = requests.get(_API_BASE, params=params, headers=_HEADERS, timeout=15)
        resp.raise_for_status()
        data = resp.json()

        if data.get("code") != 10000:
            logger.warning("七麦API返回异常: %s", data.get("msg", ""))
            return _empty_result(country, chart_type)

        for entry in data.get("rankInfo", [])[:top]:
            app_info = entry.get("appInfo", {})
            rank_b = entry.get("rank_b", {})
            rank_c = entry.get("rank_c", {})
            company = entry.get("company", {})

            # 游戏子分类
            sub_genre = ""
            if isinstance(rank_c, dict):
                sub_genre = rank_c.get("genre", "")

            # 厂商信息
            publisher = ""
            if isinstance(company, dict):
                publisher = company.get("name", "") or company.get("companyName", "")

            items.append({
                "rank":       entry.get("index", 0),
                "app_name":   app_info.get("appName", ""),
                "app_id":     app_info.get("appId", ""),
                "publisher":  publisher,
                "icon":       app_info.get("icon", ""),
                "last_rank":  _get_last_rank(rank_b),
                "change":     _get_rank_change_str(rank_b),
                "genre":      sub_genre,
                "is_ad":      entry.get("is_ad", False),
            })

    except Exception as e:
        logger.error("七麦榜单抓取失败 country=%s type=%s: %s", country, chart_type, e)
        return _empty_result(country, chart_type)

    return {
        "country":      country,
        "country_name": REGIONS.get(country, {}).get("name", country),
        "chart_type":   chart_type,
        "chart_name":   CHART_TYPES.get(chart_type, chart_type),
        "fetch_time":   datetime.now(tz=timezone.utc).isoformat(),
        "items":        items,
    }


def _empty_result(country: str, chart_type: str) -> Dict[str, Any]:
    return {
        "country":      country,
        "country_name": REGIONS.get(country, {}).get("name", country),
        "chart_type":   chart_type,
        "chart_name":   CHART_TYPES.get(chart_type, chart_type),
        "fetch_time":   datetime.now(tz=timezone.utc).isoformat(),
        "items":        [],
    }


def fetch_multi_region_ranks(
    countries: List[str] = None,
    chart_types: List[str] = None,
    top: int = 10,
) -> Dict[str, Any]:
    """
    批量抓取多个地区、多个榜单类型。

    Returns:
        {
            "cn": {
                "free":     {country, country_name, chart_type, chart_name, items},
                "grossing": {...},
            },
            "hk": {...},
            ...
            "fetch_time": "...",
            "anomalies": [  # 重点异动摘要（排名大幅变化）
                {"region": "中国大陆", "chart": "畅销榜", "app": "原神", "change": "↑22"},
                ...
            ]
        }
    """
    if not countries:
        countries = ["cn", "hk", "tw", "us", "jp", "kr"]
    if not chart_types:
        chart_types = ["free", "grossing"]

    # 生成缓存 key（按地区+类型组合）
    cache_key = ",".join(sorted(countries)) + "|" + ",".join(sorted(chart_types)) + "|" + str(top)

    # 尝试读取有效缓存
    cache = _load_cache()
    if cache_key in cache:
        entry = cache[cache_key]
        age = datetime.now(tz=timezone.utc).timestamp() - entry.get("cached_at", 0)
        if age < _CACHE_TTL:
            logger.info("七麦榜单使用缓存数据（%.0f 分钟前）", age / 60)
            return entry["data"]

    result = {
        "fetch_time": datetime.now(tz=timezone.utc).isoformat(),
        "anomalies": [],
    }

    for country in countries:
        result[country] = {}
        for ct in chart_types:
            rank_data = fetch_rank(country=country, chart_type=ct, top=top)
            result[country][ct] = rank_data

            # 检测重点异动：仅关注上升的游戏（排名升高 >= 3 位，即 rank_b.change >= 3）
            for item in rank_data["items"]:
                change_str = item.get("change", "-")
                if change_str.startswith("↑"):
                    try:
                        change_val = int(change_str[1:])
                        if change_val >= 3:
                            result["anomalies"].append({
                                "region":    rank_data["country_name"],
                                "chart":     rank_data["chart_name"],
                                "app":       item["app_name"],
                                "rank":      item["rank"],
                                "change":    change_str,
                                "last_rank": item["last_rank"],
                            })
                    except ValueError:
                        pass

    # 按变化幅度排序异动
    result["anomalies"].sort(
        key=lambda x: int(x["change"][1:]) if x["change"].startswith("↑") else 0,
        reverse=True,
    )

    # 只有数据非空才写入缓存，避免封禁时的空结果污染缓存
    total_items = sum(
        len(result.get(c, {}).get(ct, {}).get("items", []))
        for c in countries for ct in chart_types
    )
    if total_items > 0:
        cache[cache_key] = {
            "cached_at": datetime.now(tz=timezone.utc).timestamp(),
            "data": result,
        }
        _save_cache(cache)

    return result


# ==================== 每日历史榜单 + Excel 导出 ====================

def fetch_rank_by_date(
    country: str,
    chart_type: str,
    date_str: str,   # "YYYY-MM-DD"
    top: int = 30,
) -> Dict[str, Any]:
    """抓取指定日期的榜单（七麦支持 date 参数查历史）。"""
    params = {
        "country": country,
        "genre":   GENRE_GAMES,
        "device":  "iphone",
        "brand":   chart_type,   # 正确参数：brand=free/grossing（type 参数无效）
        "date":    date_str,
    }
    try:
        resp = requests.get(_API_BASE, params=params, headers=_HEADERS, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") != 10000:
            logger.warning("七麦历史API异常 %s %s %s: %s", country, chart_type, date_str, data.get("msg", ""))
            return {}
        result = {}
        for entry in data.get("rankInfo", [])[:top]:
            app_info = entry.get("appInfo", {})
            name = app_info.get("appName", "")
            rank = entry.get("index", 0)
            if name:
                result[name] = rank
        return result
    except Exception as e:
        logger.error("历史榜单抓取失败 %s %s %s: %s", country, chart_type, date_str, e)
        return {}


def fetch_weekly_rank_excel(
    countries: List[str],
    chart_types: List[str],
    top: int = 10,
) -> tuple:
    """
    抓取过去 7 天每日榜单，生成 Excel 文件。

    表格结构（参考图示）：
      标题行："{地区} {榜单类型} TOP{top}（日期范围）"
      表头行：排名 | 日期1 | 日期2 | ... | 日期7
      数据行：1 | 游戏A | 游戏A | 游戏B | ...   （每行一个排名位，每列一天的游戏名）

    同时检测异动：最后一天上榜 但 第一天未上榜 = ★新上榜；
    最后一天排名 < 第一天排名（数字更小=更靠前）= 上升。

    Returns:
        (excel_bytes, anomalies)
    """
    from datetime import date, timedelta
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    import time

    today = date.today()
    # 过去7天（含今天），从旧到新
    display_dates = [(today - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 配色（深色头像风格）──
    C_TITLE_BG   = "0F1824"   # 标题行背景
    C_TITLE_FG   = "F5B820"   # 标题行字体（金色）
    C_HEAD_BG    = "1A2840"   # 列头背景
    C_HEAD_FG    = "F5B820"   # 列头字体
    C_RANK_BG    = "0F1824"   # 排名列背景
    C_RANK_FG    = "6A8090"   # 排名列字体
    C_ODD_BG     = "141E2E"
    C_EVEN_BG    = "0F1824"
    C_RANK1_FG   = "FFD700"   # 金
    C_RANK2_FG   = "C0C0C0"   # 银
    C_RANK3_FG   = "CD7F32"   # 铜
    C_TEXT       = "C8D4E0"
    C_NEW_FG     = "A78BFA"   # 新上榜（紫）
    C_NEW_BG     = "1E1535"
    C_BORDER     = "243450"

    thin = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style(c, bg, fg, bold=False, size=10, h_align="center", wrap=False):
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(color=fg, bold=bold, size=size)
        c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
        c.border = border

    region_names = REGIONS
    all_anomalies = []
    all_rank_tables = []  # [{title, dates, rows}] for frontend Markdown rendering

    for country in countries:
        cname = region_names.get(country, {}).get("name", country)
        for chart_type in chart_types:
            chart_name = CHART_TYPES.get(chart_type, chart_type)
            sheet_name = f"{cname}_{chart_name}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_view.showGridLines = False

            # ── 抓取7天数据 ──
            # daily_data[i] = {rank(int): game_name}，对应 display_dates[i]
            daily_data = []
            for d in display_dates:
                logger.info("抓取 %s %s %s", country, chart_type, d)
                raw = fetch_rank_by_date(country, chart_type, d, top)
                # raw = {game_name: rank_int}，转为 {rank: game_name}
                rank_to_name = {int(v): k for k, v in raw.items() if v is not None}
                daily_data.append(rank_to_name)
                time.sleep(0.5)

            date_range = f"{display_dates[0][5:].replace('-','月')}日-{display_dates[-1][5:].replace('-','月')}日"
            n_cols = 1 + len(display_dates)   # 排名列 + 7天列

            # ── 第1行：大标题（跨列合并）──
            title = f"{cname} App Store 游戏{chart_name} TOP{top}（{date_range}）"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
            tc = ws.cell(row=1, column=1, value=title)
            style(tc, C_TITLE_BG, C_TITLE_FG, bold=True, size=11)
            ws.row_dimensions[1].height = 30

            # ── 第2行：列头（排名 | 月/日 × 7）──
            rc = ws.cell(row=2, column=1, value="排名")
            style(rc, C_HEAD_BG, C_HEAD_FG, bold=True)
            for ci, d in enumerate(display_dates, 2):
                # 显示"月/日"格式
                m, day = d[5:7], d[8:]
                label = f"{int(m)}月{int(day)}日"
                hc = ws.cell(row=2, column=ci, value=label)
                style(hc, C_HEAD_BG, C_HEAD_FG, bold=True)
            ws.row_dimensions[2].height = 26

            # ── 收集表格数据（供前端 Markdown 渲染）──
            date_labels = [f"{int(d[5:7])}月{int(d[8:])}日" for d in display_dates]
            table_rows = []  # list of [game_day1, game_day2, ..., game_day7]

            # ── 数据行：行=排名1~top，列=每天游戏名 ──
            for rank in range(1, top + 1):
                row = rank + 2
                is_odd = rank % 2 == 1
                row_bg = C_ODD_BG if is_odd else C_EVEN_BG

                # 排名列
                rank_fg = C_RANK1_FG if rank == 1 else C_RANK2_FG if rank == 2 else C_RANK3_FG if rank == 3 else C_RANK_FG
                nc = ws.cell(row=row, column=1, value=rank)
                style(nc, C_RANK_BG, rank_fg, bold=(rank <= 3), size=11)
                ws.row_dimensions[row].height = 22

                # 7天游戏名列
                row_games = []
                for di, day_dict in enumerate(daily_data):
                    game = day_dict.get(rank, "")
                    is_last = (di == len(daily_data) - 1)
                    is_new = is_last and game and (game not in daily_data[0].values())

                    gc = ws.cell(row=row, column=di + 2, value=game)
                    if is_new:
                        style(gc, C_NEW_BG, C_NEW_FG, bold=True, h_align="center")
                    else:
                        cell_fg = C_RANK1_FG if rank == 1 else C_RANK2_FG if rank == 2 else C_RANK3_FG if rank == 3 else C_TEXT
                        style(gc, row_bg, cell_fg, bold=(rank <= 3 and bool(game)), h_align="center")

                    row_games.append(game or "-")

                    # 异动检测：最后一天
                    if is_last and game:
                        if is_new:
                            all_anomalies.append({
                                "app": game, "region": cname, "chart": chart_name,
                                "change": "★新上榜", "rank": rank, "country": country,
                            })
                        else:
                            first_rank = next((r for r, n in daily_data[0].items() if n == game), None)
                            if first_rank and first_rank - rank >= 3:
                                all_anomalies.append({
                                    "app": game, "region": cname, "chart": chart_name,
                                    "change": f"↑{first_rank - rank}", "rank": rank, "country": country,
                                })

                table_rows.append(row_games)

            # 保存此 sheet 的表格结构
            all_rank_tables.append({
                "title": f"{cname} {chart_name} TOP{top}（{date_range}）",
                "dates": date_labels,
                "rows": table_rows,
                "country": country,  # 用于翻译判断
            })

            # ── 底部来源行 ──
            src_row = top + 3
            ws.merge_cells(start_row=src_row, start_column=1, end_row=src_row, end_column=n_cols)
            sc = ws.cell(row=src_row, column=1, value="资料来源：七麦数据")
            style(sc, C_TITLE_BG, "445566", size=9, h_align="left")
            ws.row_dimensions[src_row].height = 18

            # ── 列宽 ──
            ws.column_dimensions["A"].width = 6    # 排名列窄
            for ci in range(2, n_cols + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 16

            ws.freeze_panes = "B3"

    # ── 翻译所有 rank_tables 中的日韩游戏名 ──
    _translate_jpkr_in_tables(all_rank_tables, all_anomalies, countries)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), all_anomalies, all_rank_tables


def _translate_jpkr_in_tables(rank_tables: list, anomalies: list, countries: list):
    """
    对 rank_tables 和 anomalies 里的游戏名做日韩翻译（原地修改）。
    jp/kr 地区全翻译；其他地区含假名/韩文的也翻译。
    """
    import re as _re

    def _needs_translate(name: str, country: str) -> bool:
        if country in ("jp", "kr"):
            return True
        # 含平假名、片假名或韩文字母
        return bool(_re.search(r'[\u3040-\u30FF\uAC00-\uD7AF\u3131-\u318E]', name))

    # 收集所有需要翻译的游戏名（去重）
    to_translate = set()
    for t in rank_tables:
        c = t.get("country", "")
        for row in t.get("rows", []):
            for name in row:
                if name and name != "-" and _needs_translate(name, c):
                    to_translate.add(name)
    for a in anomalies:
        c = a.get("country", "")
        name = a.get("app", "")
        if name and _needs_translate(name, c):
            to_translate.add(name)

    if not to_translate:
        return

    try:
        from qwen_client import get_qwen_client
        import json as _json, re as _re2
        client = get_qwen_client()
        names_list = list(to_translate)

        # 分批翻译，每批最多 30 个，合并结果
        trans_map = {}
        batch_size = 30
        for i in range(0, len(names_list), batch_size):
            batch = names_list[i:i + batch_size]
            prompt = (
                "请将以下游戏名称翻译为中文（如有官方中文名优先使用官方名，否则意译）。"
                "直接返回 JSON 对象，格式：{\"原名\": \"中文名\", ...}，不要其他内容。\n\n"
                + "\n".join(f"- {n}" for n in batch)
            )
            try:
                completion = client.chat.completions.create(
                    model="qwen-turbo",
                    messages=[{"role": "user", "content": prompt}],
                    extra_body={"enable_thinking": False},
                )
                raw = (completion.choices[0].message.content or "").strip()
                fence = _re2.search(r'```(?:json)?\s*([\s\S]*?)\s*```', raw)
                if fence:
                    raw = fence.group(1).strip()
                batch_map = _json.loads(raw)
                trans_map.update(batch_map)
            except Exception as e:
                logger.warning("翻译批次 %d 失败: %s", i // batch_size, e)
    except Exception as e:
        logger.warning("rank_tables 翻译失败: %s", e)
        return

    # 回写到 rank_tables
    for t in rank_tables:
        new_rows = []
        for row in t.get("rows", []):
            new_rows.append([trans_map.get(name, name) for name in row])
        t["rows"] = new_rows

    # 回写到 anomalies
    for a in anomalies:
        a["app"] = trans_map.get(a.get("app", ""), a.get("app", ""))

