"""小红书粉丝打标模块：二维码登录 + Playwright UI驱动抓取 + AI 打标 + 导出 Excel。"""

import asyncio
import json
import logging
import time
import base64
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).parent.parent / "fan_tag_output"
OUTPUT_DIR.mkdir(exist_ok=True)
CACHE_FILE = OUTPUT_DIR / "progress_cache.json"
PROFILE_DIR = Path(__file__).parent.parent / "xhs_chrome_profile"

_task_state: Dict[str, Any] = {}


def get_task_state(task_id: str) -> Dict[str, Any]:
    return _task_state.get(task_id, {})


def _set_state(task_id: str, **kwargs):
    _task_state.setdefault(task_id, {}).update(kwargs)


def _log(task_id: str, msg: str):
    state = _task_state.setdefault(task_id, {})
    state.setdefault("logs", []).append(msg)
    logger.info("[fan_tag:%s] %s", task_id, msg)


# ─── XHS Browser Session ─────────────────────────────────────────────────────
class XhsSession:
    HOME = "https://www.xiaohongshu.com"
    EDITH = "https://edith.xiaohongshu.com"

    def __init__(self):
        self._pw = None
        self._ctx = None
        self._page = None
        self._captured: Dict[str, Any] = {}   # path -> response data
        self._all_captured: List[Dict] = []    # (url, body) 所有 XHS 接口响应
        self._logged_in = False

    async def start(self):
        from playwright.async_api import async_playwright
        self._pw = await async_playwright().start()
        PROFILE_DIR.mkdir(exist_ok=True)
        # 在启动前清除 Session Restore 状态，防止旧标签（创作服务平台等）被恢复
        last_session = PROFILE_DIR / "Default" / "Last Session"
        last_tabs   = PROFILE_DIR / "Default" / "Last Tabs"
        for f in (last_session, last_tabs):
            try:
                if f.exists():
                    f.unlink()
            except Exception:
                pass

        self._ctx = await self._pw.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            channel="chrome",
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--profile-directory=Default",
                "--window-size=1280,900",
                "--restore-last-session=false",
            ],
            timeout=30000,
        )
        self._page = await self._ctx.new_page()

        # 拦截所有 XHS API 响应
        async def _on_resp(resp):
            url = resp.url
            if "xiaohongshu.com/api" in url or "edith.xiaohongshu.com" in url:
                try:
                    body = await resp.json()
                    path = re.sub(r'^https?://[^/]+', '', url).split("?")[0]
                    qs = url.split("?", 1)[1] if "?" in url else ""
                    if isinstance(body, dict):
                        self._captured[path] = body
                        if qs:
                            self._captured[path + "?" + qs] = body
                    # 保存全量列表（包含 list 类型，由 _scan_captured 处理）
                    self._all_captured.append({"url": url, "path": path, "body": body})
                except Exception:
                    pass
        self._ctx.on("response", _on_resp)

    async def close(self):
        if self._ctx:
            try:
                await self._ctx.close()
            except Exception:
                pass
        if self._pw:
            try:
                await self._pw.stop()
            except Exception:
                pass

    def _get_data(self, path: str) -> Optional[Dict]:
        body = self._captured.get(path)
        if not isinstance(body, dict):
            return None
        if body.get("success"):
            data = body.get("data")
            return data if isinstance(data, dict) else None
        return None

    async def check_login(self) -> bool:
        """访问首页检查是否已登录。"""
        self._captured.clear()
        self._all_captured.clear()
        await self._page.goto(self.HOME + "/explore", wait_until="domcontentloaded", timeout=25000)
        await asyncio.sleep(3)
        me = self._get_data("/api/sns/web/v2/user/me")
        if me and not me.get("guest"):
            self._logged_in = True
            return True
        return False

    async def get_qr_screenshot(self) -> Optional[str]:
        """返回登录页二维码截图的 base64。"""
        await self._page.goto(self.HOME + "/login", wait_until="domcontentloaded", timeout=25000)
        await asyncio.sleep(3)
        try:
            await self._page.wait_for_selector("canvas, img[src*='qr'], .qrcode, [class*='qr']", timeout=5000)
        except Exception:
            pass
        png = await self._page.screenshot(type="png", full_page=False)
        return base64.b64encode(png).decode()

    async def wait_login(self, timeout: float = 120) -> bool:
        """等待用户扫码登录，轮询检测。"""
        deadline = asyncio.get_event_loop().time() + timeout
        while asyncio.get_event_loop().time() < deadline:
            await asyncio.sleep(3)
            url = self._page.url
            if "/login" not in url and "/captcha" not in url:
                self._captured.clear()
                self._all_captured.clear()
                await self._page.goto(self.HOME + "/explore", wait_until="domcontentloaded", timeout=20000)
                await asyncio.sleep(2)
                me = self._get_data("/api/sns/web/v2/user/me")
                if me and not me.get("guest"):
                    self._logged_in = True
                    return True
        return False

    # ── 核心：被动监听模式获取粉丝列表 ──────────────────────────────────────
    async def fetch_followers_by_ui(self, task_id: str, target_user_id: str) -> List[str]:
        """
        被动监听模式：脚本不主动导航（避免被验证码拦截），
        等待用户在 Chrome 窗口中手动打开粉丝页并滚动，
        脚本自动拦截 API + 提取 DOM 中的用户 ID。
        """
        ids: List[str] = []
        seen: set = set()

        def _scan_captured():
            """扫描所有已拦截响应，提取 user_id。"""
            added = 0
            for item in self._all_captured:
                body = item["body"]
                if not body:
                    continue
                # body 可能本身就是 list，也可能是 dict
                if isinstance(body, list):
                    candidates = body
                elif isinstance(body, dict):
                    if not body.get("success"):
                        continue
                    data = body.get("data") or {}
                    if isinstance(data, list):
                        candidates = data
                    elif isinstance(data, dict):
                        candidates = (data.get("users") or data.get("followers") or
                                      data.get("fans") or data.get("list") or
                                      data.get("notifications") or [])
                    else:
                        continue
                else:
                    continue

                for u in (candidates if isinstance(candidates, list) else []):
                    if not isinstance(u, dict):
                        continue
                    user_obj = u.get("user")
                    uid = str(
                        u.get("user_id") or u.get("userid") or
                        u.get("userId") or u.get("id") or
                        (user_obj.get("user_id") if isinstance(user_obj, dict) else "") or ""
                    )
                    if uid and len(uid) > 8 and uid not in seen and uid != target_user_id:
                        seen.add(uid)
                        ids.append(uid)
                        added += 1
            return added

        async def _scan_dom():
            """扫描当前页面 DOM 中的用户链接。"""
            added = 0
            try:
                dom_ids = await self._page.evaluate("""() => {
                    const ids = new Set();
                    document.querySelectorAll('a[href*="/user/profile/"]').forEach(a => {
                        const m = (a.href || '').match(/\\/user\\/profile\\/([0-9a-f]{16,})/i);
                        if (m) ids.add(m[1]);
                    });
                    return [...ids];
                }""")
                for uid in (dom_ids or []):
                    uid = str(uid).strip()
                    if uid and len(uid) > 8 and uid not in seen and uid != target_user_id:
                        seen.add(uid)
                        ids.append(uid)
                        added += 1
            except Exception:
                pass
            return added

        # ── 1. 直接导航到小红书通知页（新增关注），被动监听 ──────────────
        _log(task_id, "🌐 打开小红书通知页（新增关注）...")
        self._captured.clear()
        self._all_captured.clear()

        try:
            await self._page.goto(
                "https://www.xiaohongshu.com/notification",
                wait_until="domcontentloaded", timeout=25000
            )
            await asyncio.sleep(4)
            _log(task_id, f"📍 当前: {self._page.url}")
        except Exception as e:
            _log(task_id, f"⚠️ 导航失败: {e}")

        # Session Restore 完成后关闭所有其他标签（创作服务平台等旧页面）
        await asyncio.sleep(3)
        for p in list(self._ctx.pages):
            if p != self._page:
                try:
                    _log(task_id, f"🗑️ 关闭旧标签: {p.url[:80]}")
                    await p.close()
                except Exception:
                    pass
        try:
            await self._page.bring_to_front()
        except Exception:
            pass

        # ── 2. 提示用户在通知页点击「新增关注」并滚动 ────────────────────
        _log(task_id, "👇 请在 Chrome 窗口里点击「新增关注」，然后慢慢向下滚动...")
        _set_state(task_id,
                   stage="fetch_list",
                   message="请在 Chrome 窗口中：①点击「新增关注」选项卡 ②慢慢向下滚动至底部 ③滚完后点击「完成」按钮",
                   need_manual=True,
                   manual_done=False)

        # 被动监听循环，最多等 10 分钟
        max_wait = 600
        elapsed = 0
        interval = 3

        while elapsed < max_wait:
            await asyncio.sleep(interval)
            elapsed += interval

            state = _task_state.get(task_id, {})
            if state.get("manual_done"):
                _log(task_id, "✅ 用户确认操作完毕，停止监听")
                break

            api_added = _scan_captured()
            dom_added = await _scan_dom()
            total_now = len(ids)

            if api_added + dom_added > 0:
                _log(task_id, f"📡 新增 {api_added+dom_added} 人，当前共 {total_now} 人")

            _set_state(task_id,
                       follower_count=total_now,
                       message=f"监听中... 已发现 {total_now} 位粉丝（继续滚动后点击「完成」）")

        return ids

    # ── 获取单个用户信息（浏览器原生请求）────────────────────────────────────
    async def fetch_user_info_by_page(self, user_id: str) -> Dict[str, Any]:
        """访问用户主页，从 DOM 直接抓取昵称、简介、粉丝数等信息。"""
        self._captured.clear()
        self._all_captured.clear()
        await self._page.goto(
            f"{self.HOME}/user/profile/{user_id}",
            wait_until="domcontentloaded",
            timeout=25000,
        )
        await asyncio.sleep(3)

        # 从 DOM 抓取用户信息
        try:
            info = await self._page.evaluate("""() => {
                try {
                    // 昵称
                    const nickEl = document.querySelector(
                        '.user-name, [class*="username"], [class*="user-name"], h1[class*="user"]'
                    );
                    const nick = (nickEl && nickEl.innerText.trim()) || '';

                    // 简介
                    const bioEl = document.querySelector(
                        '[class*="user-desc"], [class*="desc"][class*="user"], .desc'
                    );
                    const bio = (bioEl && bioEl.innerText.trim()) || '';

                    // 数量（关注/粉丝/笔记）
                    const countEls = [...document.querySelectorAll('[class*="count"]')]
                        .filter(el => el.innerText.trim());
                    const counts = countEls.slice(0, 4).map(el => el.innerText.trim());

                    // 笔记标题
                    const noteEls = [...document.querySelectorAll(
                        '[class*="note"] [class*="title"], [class*="feeds"] [class*="title"]'
                    )].slice(0, 5);
                    const noteTitles = noteEls.map(el => el.innerText.trim()).filter(t => t);

                    return {nick, bio, counts, noteTitles};
                } catch(e) { return null; }
            }""")
            if info and info.get('nick'):
                # counts 顺序通常是 关注数、粉丝数、笔记数
                counts = info.get('counts', [])
                def to_int(s):
                    try:
                        s = str(s).replace(',','').replace('万','0000').strip()
                        return int(float(s))
                    except:
                        return 0
                return {
                    "basic_info": {
                        "nickname": info.get('nick', ''),
                        "desc": info.get('bio', ''),
                        "imageb": "",
                    },
                    "interactions": [
                        {"type": "follows", "count": to_int(counts[0]) if len(counts) > 0 else 0},
                        {"type": "fans",    "count": to_int(counts[1]) if len(counts) > 1 else 0},
                        {"type": "notes",   "count": to_int(counts[2]) if len(counts) > 2 else 0},
                    ],
                    "_note_titles": info.get('noteTitles', []),
                }
        except Exception:
            pass

        return {}

    async def fetch_user_notes_by_page(self, user_id: str) -> List[Dict]:
        """从已加载的页面 DOM 中提取笔记标题。"""
        try:
            # 优先用 _note_titles（fetch_user_info_by_page 已抓到的）
            info = await self._page.evaluate("""() => {
                try {
                    const noteEls = [...document.querySelectorAll(
                        '[class*="note"] [class*="title"], [class*="feeds"] [class*="title"], .note-info .title'
                    )].slice(0, 5);
                    return noteEls.map(el => ({title: el.innerText.trim()})).filter(n => n.title);
                } catch(e) { return []; }
            }""")
            return info or []
        except Exception:
            return []


# ─── 解析粉丝详情 ─────────────────────────────────────────────────────────────
def _parse_fan_detail(user_id: str, data: Dict, notes: List[Dict]) -> Dict[str, Any]:
    info = {
        "user_id": user_id,
        "nickname": "",
        "bio": "",
        "avatar_url": "",
        "fans_count": 0,
        "following_count": 0,
        "notes_count": 0,
        "recent_notes": notes,
        "error": None,
    }
    if not data:
        info["error"] = "获取失败"
        return info

    basic = data.get("basic_info") or data
    info["nickname"] = basic.get("nickname") or basic.get("name") or ""
    info["bio"] = basic.get("desc") or basic.get("bio") or ""
    info["avatar_url"] = basic.get("imageb") or basic.get("images") or ""

    for item in (data.get("interactions") or []):
        t = item.get("type", "")
        try:
            count = int(str(item.get("count") or "0").replace(",", ""))
        except Exception:
            count = 0
        if t == "fans":
            info["fans_count"] = count
        elif t == "follows":
            info["following_count"] = count
        elif t == "notes":
            info["notes_count"] = count

    return info


# ─── AI 批量打标 ──────────────────────────────────────────────────────────────
_TAG_PROMPT = """你是一个小红书运营分析师，分析游戏资讯自媒体（游戏新闻/行业动态/游戏评测）的粉丝画像。
根据粉丝昵称、简介、笔记，输出 JSON 数组，字段：
- user_id（原样返回）
- identity: 游戏从业者/游戏博主UP主/游戏测评者/互联网人/学生党/普通游戏用户/无法判断
- game_pref: 手游玩家/主机PC玩家/二次元爱好者/女性向游戏/泛娱乐/不明显
- relevance: 高-核心受众/中-游戏爱好者/低-轻度相关/无关
- activity: 高/中/低-僵尸粉
- potential: 高/中/低
- reason: 一句话（15字内）
只输出 JSON 数组。"""


def _ai_tag_batch(fans: List[Dict]) -> List[Dict]:
    from qwen_client import get_qwen_client
    lines = []
    for i, f in enumerate(fans, 1):
        notes = " | ".join(n["title"] for n in f.get("recent_notes", []) if n.get("title")) or "无"
        lines.append(f"{i}. user_id={f['user_id']} | 昵称：{f['nickname'] or '无'} | "
                     f"简介：{f['bio'] or '无'} | 粉丝数：{f['fans_count']} | "
                     f"笔记数：{f['notes_count']} | 近期笔记：{notes}")
    try:
        client = get_qwen_client()
        resp = client.chat.completions.create(
            model="qwen-max",
            messages=[{"role": "system", "content": _TAG_PROMPT},
                      {"role": "user", "content": "\n".join(lines)}],
            extra_body={"enable_thinking": False},
        )
        raw = resp.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```", 2)[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.rsplit("```", 1)[0].strip()
        tags = json.loads(raw)
        if isinstance(tags, list):
            return tags
    except Exception as e:
        logger.warning("AI打标失败: %s", e)
    return [{"user_id": f["user_id"], "identity": "无法判断", "game_pref": "不明显",
             "relevance": "低-轻度相关", "activity": "低", "potential": "低", "reason": "AI打标失败"}
            for f in fans]


# ─── 导出 Excel ───────────────────────────────────────────────────────────────
def _export_excel(fans: List[Dict]) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "粉丝打标"
    headers = ["序号", "小红书ID", "昵称", "个人简介", "粉丝数", "关注数", "笔记数",
               "最近笔记标题", "身份标签", "游戏偏好", "与账号相关度", "活跃度", "互动潜力", "AI判断理由"]
    hfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wa = Alignment(vertical="top", wrap_text=True)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = ca
    fills = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
             PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")]
    for i, fan in enumerate(fans, 1):
        tags = fan.get("tags") or {}
        notes_text = "\n".join(n["title"] for n in fan.get("recent_notes", []) if n.get("title"))
        row = [i, fan.get("user_id",""), fan.get("nickname",""), fan.get("bio",""),
               fan.get("fans_count",0), fan.get("following_count",0), fan.get("notes_count",0),
               notes_text, tags.get("identity",""), tags.get("game_pref",""),
               tags.get("relevance",""), tags.get("activity",""), tags.get("potential",""), tags.get("reason","")]
        fill = fills[(i-1)%2]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i+1, column=ci, value=val)
            c.fill = fill; c.alignment = wa if ci in (4, 8) else ca
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for ci, w in enumerate([6, 26, 18, 30, 10, 10, 10, 35, 18, 16, 16, 10, 10, 24], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    for i in range(2, len(fans)+2):
        ws.row_dimensions[i].height = 30
    path = OUTPUT_DIR / f"粉丝打标_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(str(path))
    return path


# ─── 缓存 ─────────────────────────────────────────────────────────────────────
def _save_cache(data: Dict):
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning("缓存保存失败: %s", e)


# ─── 主流程 ───────────────────────────────────────────────────────────────────
async def _run_async(task_id: str, cookie: str, target_user_id: str):
    _set_state(task_id, status="running", stage="init", message="启动浏览器...",
               total=0, done=0, current_name="", excel_path=None, error=None,
               logs=[], qr_image=None, need_login=False)

    # 加载缓存
    cache: Dict[str, Any] = {}
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, encoding="utf-8") as f:
                cache = json.load(f)
            _log(task_id, f"✅ 加载断点缓存，已有 {len(cache.get('fans', {}))} 条数据")
        except Exception:
            pass

    session = XhsSession()
    try:
        _log(task_id, "🌐 启动 Chrome 浏览器...")
        await session.start()
        _log(task_id, "✅ 浏览器已启动")

        # 检查登录态
        logged_in = await session.check_login()
        if not logged_in:
            _log(task_id, "⚠️ 未登录，正在显示二维码...")
            qr_b64 = await session.get_qr_screenshot()
            _set_state(task_id, need_login=True, qr_image=qr_b64,
                       message="请用小红书 App 扫描二维码登录", stage="login")
            logged_in = await session.wait_login(timeout=180)
            if not logged_in:
                raise RuntimeError("等待扫码超时（3分钟），请重新开始")
            _set_state(task_id, need_login=False, qr_image=None)

        _log(task_id, "✅ 已登录小红书！")

        # ── Step 1：获取粉丝列表 ────────────────────────────────────────────
        cached_ids = cache.get("follower_ids")
        if cached_ids and len(cached_ids) >= 30:
            follower_ids = cached_ids
            _log(task_id, f"✅ 使用缓存粉丝列表，共 {len(follower_ids)} 人")
        else:
            _set_state(task_id, stage="fetch_list", message="获取粉丝列表中...")
            _log(task_id, "🔍 开始获取粉丝列表（UI 驱动模式）...")
            follower_ids = await session.fetch_followers_by_ui(task_id, target_user_id)
            cache["follower_ids"] = follower_ids
            cache["task_id"] = task_id
            _save_cache(cache)
            _log(task_id, f"✅ 粉丝列表获取完成，共 {len(follower_ids)} 人")

        if not follower_ids:
            raise RuntimeError(
                "未获取到粉丝 ID。请重试：点击「清除缓存」后重新启动，"
                "然后在弹出的 Chrome 窗口中手动打开粉丝页并滚动。"
            )

        total = len(follower_ids)
        _set_state(task_id, total=total, stage="fetch_detail")
        _log(task_id, f"📋 共 {total} 位粉丝，开始抓取详情...")

        # ── Step 2：逐人抓取详情 ────────────────────────────────────────────
        fans_data: Dict[str, Any] = cache.get("fans", {})
        new_fetched = 0

        for idx, fan_id in enumerate(follower_ids, 1):
            if fan_id in fans_data:
                _set_state(task_id, done=idx,
                           current_name=fans_data[fan_id].get("nickname", fan_id))
                continue

            _set_state(task_id, done=idx, message=f"抓取 {idx}/{total}: {fan_id}")

            user_data = await session.fetch_user_info_by_page(fan_id)
            notes = await session.fetch_user_notes_by_page(fan_id) if user_data else []

            fan_info = _parse_fan_detail(fan_id, user_data, notes)
            fans_data[fan_id] = fan_info
            new_fetched += 1

            name = fan_info.get("nickname") or fan_id
            _set_state(task_id, current_name=name,
                       message=f"已抓取 {idx}/{total}：{name}")

            if new_fetched % 10 == 0:
                cache["fans"] = fans_data
                _save_cache(cache)
                _log(task_id, f"💾 进度已保存 ({idx}/{total})")

            await asyncio.sleep(1.5)

        cache["fans"] = fans_data
        _save_cache(cache)
        _log(task_id, f"✅ 详情抓取完成，共 {len(fans_data)} 人")

    finally:
        await session.close()

    # ── Step 3：AI 打标 ─────────────────────────────────────────────────────
    _set_state(task_id, stage="ai_tag", message="AI 批量打标中...")
    _log(task_id, "🤖 开始 AI 打标...")

    fans_list = [fans_data[uid] for uid in follower_ids if uid in fans_data]
    tagged: Dict[str, Any] = cache.get("tags", {})
    batches = [fans_list[i:i+20] for i in range(0, len(fans_list), 20)]

    for bi, batch in enumerate(batches, 1):
        untagged = [f for f in batch if f["user_id"] not in tagged]
        if not untagged:
            continue
        _set_state(task_id, message=f"AI 打标 {bi}/{len(batches)} 批...")
        _log(task_id, f"🤖 第 {bi}/{len(batches)} 批，{len(untagged)} 人")
        for tag in _ai_tag_batch(untagged):
            uid = str(tag.get("user_id", ""))
            if uid:
                tagged[uid] = tag
        cache["tags"] = tagged
        _save_cache(cache)
        await asyncio.sleep(1)

    _log(task_id, f"✅ AI 打标完成，共 {len(tagged)} 人")
    for f in fans_list:
        f["tags"] = tagged.get(f["user_id"], {})

    # ── Step 4：导出 Excel ──────────────────────────────────────────────────
    _set_state(task_id, stage="export", message="导出 Excel...")
    _log(task_id, "📊 导出 Excel...")
    excel_path = _export_excel(fans_list)
    _log(task_id, f"✅ 完成！Excel: {excel_path.name}")

    if CACHE_FILE.exists():
        CACHE_FILE.unlink()

    _set_state(task_id, status="done", stage="done", message="全部完成！",
               excel_path=str(excel_path), done=total)


def run_fan_tag_task(task_id: str, cookie: str, target_user_id: str):
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(_run_async(task_id, cookie, target_user_id))
    except Exception as e:
        logger.exception("fan_tag 异常: %s", e)
        _set_state(task_id, status="error", error=str(e), message=f"出错: {e}")
    finally:
        try:
            loop.close()
        except Exception:
            pass
