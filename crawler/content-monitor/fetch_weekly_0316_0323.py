#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
七麦数据历史榜单爬取 & 翻译脚本
爬取 2026-03-16 ~ 2026-03-23（8天）各地区 App Store 游戏免费榜 + 畅销榜 TOP30
并将所有非中文游戏名翻译为中文

输出：
  weekly_cache/qimai_raw_0316_0323.json   -- 原始JSON数据（支持断点续爬）
  weekly_cache/translations_0316_0323.json -- 翻译缓存
  weekly_cache/七麦榜单_0316-0323.xlsx    -- Excel报表（深色主题，含翻译）

用法：
  cd D:/project/content-monitor
  python fetch_weekly_0316_0323.py
"""

import json
import logging
import os
import re
import sys
import time
import io
from datetime import date, timedelta
from pathlib import Path
from typing import Dict, List, Set

import requests

# 强制 stdout/stderr 使用 UTF-8（解决 Windows GBK 控制台乱码）
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ─────────────────────────── 日志 ────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

# ─────────────────────────── 路径配置 ────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
OUTPUT_DIR   = SCRIPT_DIR / "weekly_cache"
OUTPUT_DIR.mkdir(exist_ok=True)

RAW_JSON_PATH   = OUTPUT_DIR / "qimai_raw_0316_0323.json"
TRANS_JSON_PATH = OUTPUT_DIR / "translations_0316_0323.json"
EXCEL_PATH      = OUTPUT_DIR / "七麦榜单_0316-0323.xlsx"

# ─────────────────────────── 日期范围 ────────────────────────────────────────
START_DATE = date(2026, 3, 16)
END_DATE   = date(2026, 3, 23)

DISPLAY_DATES: List[str] = []
_d = START_DATE
while _d <= END_DATE:
    DISPLAY_DATES.append(_d.strftime("%Y-%m-%d"))
    _d += timedelta(days=1)
# → ['2026-03-16', ..., '2026-03-23']  共8天

# ─────────────────────────── API 配置 ────────────────────────────────────────
_API_BASE = "https://api.qimai.cn/rank/index"

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.qimai.cn/",
    "Cookie": (
        "qm_check=A1sdRUIQChtxen8pI0dAMRcOUFseEHBeQF0JTjVBWDgIVh9mfRAQZEoKRiAaGQkSXFRBEBEATAhVVldfPExPHBd2WUtVV0xWIlZZWA8JagJtABlAR2dQOVdEWD1JcQYDGRscFlsNAwQDWkNYBRsCHAkcBBRVRBo%3D; "
        "PHPSESSID=5be4jvic5bi221ar6lcdk7cnb6; "
        "gr_user_id=c84caf3d-9846-4f2e-be8e-547a56d0155e; "
        "ada35577182650f1_gr_session_id=ded42ff1-2f50-421b-8646-ab654a41bcbc; "
        "ada35577182650f1_gr_session_id_sent_vst=ded42ff1-2f50-421b-8646-ab654a41bcbc; "
        "USERINFO=RrS%2BfEQ1ijCHAV55XHZfN%2FMUn2dFNZU%2BicQFas0TpyTxDDlpN%2BScxIFPKZ9ooNEsTGOPBB4NTLt1crFniSlAMdzOQO0%2BnplvW0Iyb50SHnUw%2F201XxjWOSGpSlGVYnVtf5FX5R0tjw5vC1bLhSl8Dw%3D%3D; "
        "AUTHKEY=CxxVpjQzCRlDhA0dru%2BQzHuf0c9LEV8W%2FNe5eSAM6OUI3rVbbumOQ5o2q8tKYkAtb4kdY8Qz0HcnlbeOTxYassgbIxFd9DAUFj8fG8xDyzBW%2Byeu%2FNahEg%3D%3D; "
        "ada35577182650f1_gr_last_sent_sid_with_cs1=ded42ff1-2f50-421b-8646-ab654a41bcbc; "
        "ada35577182650f1_gr_last_sent_cs1=qm20425157953; "
        "ada35577182650f1_gr_cs1=qm20425157953; "
        "aso_ucenter=b5dfUkeQvi9XmpE13Ve9qaDVh3uzkES%2FPbplFGoXIN1H6z88jViXc1f2Gp%2FfDU9ikbA; "
        "synct=1773049528.878; syncd=-219"
    ),
}

# ─────────────────────────── 地区 & 榜单 ─────────────────────────────────────
REGIONS: Dict[str, Dict[str, str]] = {
    "cn": {"name": "中国大陆", "flag": "🇨🇳"},
    "hk": {"name": "中国香港", "flag": "🇭🇰"},
    "tw": {"name": "中国台湾", "flag": "🇹🇼"},
    "us": {"name": "美国",     "flag": "🇺🇸"},
    "jp": {"name": "日本",     "flag": "🇯🇵"},
    "kr": {"name": "韩国",     "flag": "🇰🇷"},
    "gb": {"name": "英国",     "flag": "🇬🇧"},
    "de": {"name": "德国",     "flag": "🇩🇪"},
    "fr": {"name": "法国",     "flag": "🇫🇷"},
    "sg": {"name": "新加坡",   "flag": "🇸🇬"},
    "th": {"name": "泰国",     "flag": "🇹🇭"},
    "sa": {"name": "沙特",     "flag": "🇸🇦"},
    "tr": {"name": "土耳其",   "flag": "🇹🇷"},
    "br": {"name": "巴西",     "flag": "🇧🇷"},
    "in": {"name": "印度",     "flag": "🇮🇳"},
}

CHART_TYPES: Dict[str, str] = {
    "free":     "免费榜",
    "grossing": "畅销榜",
}

GENRE_GAMES = 6014  # iOS App Store 游戏类目
TOP = 10            # 每榜前 10 名


# ═════════════════════════════════════════════════════════════════════════════
# 一、数据爬取
# ═════════════════════════════════════════════════════════════════════════════

def fetch_rank_by_date(country: str, chart_type: str, date_str: str) -> Dict[int, str]:
    """
    调用七麦 API 获取指定地区+类型+日期的历史榜单。
    返回 {排名(int): 游戏名(str)}，失败返回 {}。
    """
    params = {
        "country": country,
        "genre":   GENRE_GAMES,
        "device":  "iphone",
        "brand":   chart_type,   # brand=free / grossing（不是 type）
        "date":    date_str,
    }
    try:
        resp = requests.get(_API_BASE, params=params, headers=_HEADERS, timeout=15)
        resp.raise_for_status()
        data = resp.json()

        if data.get("code") != 10000:
            logger.warning("API异常 %s %s %s → code=%s msg=%s",
                           country, chart_type, date_str,
                           data.get("code"), data.get("msg", ""))
            return {}

        result: Dict[int, str] = {}
        for entry in data.get("rankInfo", [])[:TOP]:
            app_info = entry.get("appInfo", {})
            name = app_info.get("appName", "").strip()
            rank = entry.get("index", 0)
            if name and rank:
                result[rank] = name
        return result

    except Exception as e:
        logger.error("抓取失败 %s %s %s: %s", country, chart_type, date_str, e)
        return {}


def _load_json(path: Path) -> dict:
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_json(path: Path, data: dict):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def fetch_all_data(countries: List[str] = None, chart_types: List[str] = None) -> dict:
    """
    批量爬取所有地区×榜单类型×8天的数据，支持断点续爬。

    存储结构（raw_data）：
      raw_data[country][chart_type][date_str] = {"1": "游戏A", "2": "游戏B", ...}
    """
    if countries is None:
        countries = list(REGIONS.keys())
    if chart_types is None:
        chart_types = list(CHART_TYPES.keys())

    raw_data = _load_json(RAW_JSON_PATH)
    total = len(countries) * len(chart_types) * len(DISPLAY_DATES)
    done  = 0

    for country in countries:
        raw_data.setdefault(country, {})
        cname = REGIONS[country]["name"]
        for ct in chart_types:
            raw_data[country].setdefault(ct, {})
            chart_name = CHART_TYPES[ct]
            for d in DISPLAY_DATES:
                done += 1
                # 已有数据则跳过（断点续爬）
                if d in raw_data[country][ct] and raw_data[country][ct][d]:
                    logger.info("[%3d/%d] [skip] %s %s %s",
                                done, total, cname, chart_name, d)
                    continue

                logger.info("[%3d/%d] [fetch] %s(%s) %s %s",
                            done, total, cname, country, chart_name, d)

                result = fetch_rank_by_date(country, ct, d)
                # JSON key 必须是字符串
                raw_data[country][ct][d] = {str(k): v for k, v in result.items()}

                if result:
                    _save_json(RAW_JSON_PATH, raw_data)   # 每次成功后即时保存
                else:
                    logger.warning("  [warn] 返回空数据（可能被限频或 Cookie 失效）")

                time.sleep(0.6)   # 友好爬取，避免触发限流

    _save_json(RAW_JSON_PATH, raw_data)
    logger.info("[done] 数据爬取完毕，已保存: %s", RAW_JSON_PATH)
    return raw_data


# ═════════════════════════════════════════════════════════════════════════════
# 二、翻译
# ═════════════════════════════════════════════════════════════════════════════

def is_chinese_name(name: str) -> bool:
    """
    判断游戏名是否已经是中文（无需翻译）。
    规则：
      - 含平假名/片假名/韩文 → 非中文
      - 中文字符占比 >= 60%  → 中文
      - 其余（英文、混合等）  → 非中文
    """
    if not name or name in ("-", ""):
        return True
    # 日文假名或韩文 → 直接判为非中文
    if re.search(r"[\u3040-\u30FF\uAC00-\uD7AF\u3131-\u318E]", name):
        return False
    # 去掉空格/数字/标点后计算比例
    stripped = re.sub(r"[\s\d\W]+", "", name)
    if not stripped:
        return True
    chinese = len(re.findall(r"[\u4e00-\u9fff\u3400-\u4dbf]", stripped))
    return chinese / len(stripped) >= 0.6


def _get_qwen_client():
    """获取通义千问客户端（优先复用项目现有实现）。"""
    try:
        sys.path.insert(0, str(SCRIPT_DIR))
        from qwen_client import get_qwen_client
        return get_qwen_client()
    except ImportError:
        from openai import OpenAI
        return OpenAI(
            api_key=os.getenv("DASHSCOPE_API_KEY", ""),
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
            timeout=60.0,
        )


def translate_names(names: List[str]) -> Dict[str, str]:
    """
    批量将非中文游戏名翻译为中文，每批最多 30 个。
    支持翻译缓存（二次运行直接复用，不重复消耗 API）。
    返回 {原名: 中文名} 的完整映射字典。
    """
    trans_map = _load_json(TRANS_JSON_PATH)

    remaining = [n for n in names if n and n not in trans_map]
    if not remaining:
        logger.info("所有名称已在翻译缓存中，无需调用 API")
        return trans_map

    logger.info("待翻译 %d 个游戏名（总去重 %d 个）", len(remaining), len(names))

    try:
        client = _get_qwen_client()
    except Exception as e:
        logger.error("无法初始化 Qwen 客户端: %s", e)
        return trans_map

    BATCH = 30
    total_batches = (len(remaining) + BATCH - 1) // BATCH

    for bi, i in enumerate(range(0, len(remaining), BATCH), 1):
        batch = remaining[i: i + BATCH]
        logger.info("  翻译第 %d/%d 批（%d 个）...", bi, total_batches, len(batch))

        prompt = (
            "请将以下游戏名称翻译/转换为中文（有官方中文名优先官方名，否则意译或音译）。\n"
            "纯中文名直接原样返回，不要修改。\n"
            "只返回 JSON 对象，格式：{\"原名\": \"中文名\", ...}，不要任何其他内容。\n\n"
            + "\n".join(f"- {n}" for n in batch)
        )
        try:
            resp = client.chat.completions.create(
                model="qwen-turbo",
                messages=[{"role": "user", "content": prompt}],
            )
            raw = (resp.choices[0].message.content or "").strip()
            # 去掉 ```json ... ``` 代码块
            m = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", raw)
            if m:
                raw = m.group(1).strip()
            batch_map: Dict[str, str] = json.loads(raw)
            trans_map.update(batch_map)
            _save_json(TRANS_JSON_PATH, trans_map)
            logger.info("  [ok] 完成 %d 个", len(batch_map))
        except Exception as e:
            logger.warning("  [fail] 批次 %d 翻译失败: %s", bi, e)

        time.sleep(1.2)   # 避免 API 限流

    logger.info("[done] 翻译完成，共 %d 条翻译记录", len(trans_map))
    return trans_map


def collect_non_chinese_names(raw_data: dict) -> List[str]:
    """从原始数据中收集所有需要翻译的游戏名（去重）。"""
    seen: Set[str] = set()
    for country, charts in raw_data.items():
        if country not in REGIONS:
            continue
        for ct, dates in charts.items():
            for d, rank_map in dates.items():
                for rank_str, name in rank_map.items():
                    if name and name not in seen and not is_chinese_name(name):
                        seen.add(name)
    return sorted(seen)


# ═════════════════════════════════════════════════════════════════════════════
# 三、生成 Excel
# ═════════════════════════════════════════════════════════════════════════════

def generate_excel(raw_data: dict, trans_map: Dict[str, str]):
    """
    生成深色主题 Excel 报表（与项目 fetch_weekly_rank_excel 风格一致）。
    共 14地区 × 2榜单 = 最多 28 个 Sheet。
    新上榜游戏标注 ★，已翻译游戏用绿色区分。
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("缺少 openpyxl，请运行: pip install openpyxl")
        sys.exit(1)

    # ── 配色（深色主题，与项目一致）────────────────────────────────────────
    C_TITLE_BG = "0F1824"
    C_TITLE_FG = "F5B820"   # 金色标题
    C_HEAD_BG  = "1A2840"
    C_HEAD_FG  = "F5B820"
    C_RANK_BG  = "0F1824"
    C_RANK_FG  = "6A8090"
    C_ODD_BG   = "141E2E"
    C_EVEN_BG  = "0F1824"
    C_RANK1_FG = "FFD700"   # 金
    C_RANK2_FG = "C0C0C0"   # 银
    C_RANK3_FG = "CD7F32"   # 铜
    C_TEXT     = "C8D4E0"
    C_NEW_FG   = "A78BFA"   # 新上榜（紫色）
    C_NEW_BG   = "1E1535"
    C_TRANS_FG = "6DD98C"   # 已翻译（绿色）
    C_BORDER   = "243450"

    thin   = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style(cell, bg, fg, bold=False, size=10, h_align="center"):
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(color=fg, bold=bold, size=size)
        cell.alignment = Alignment(horizontal=h_align, vertical="center")
        cell.border    = border

    # ── 日期列头（"3月16日" … "3月23日"）─────────────────────────────────
    date_labels = [
        f"{int(d[5:7])}月{int(d[8:])}日"
        for d in DISPLAY_DATES
    ]
    n_dates = len(DISPLAY_DATES)   # 8
    n_cols  = 1 + n_dates          # 排名列 + 8天

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for country, region_info in REGIONS.items():
        if country not in raw_data:
            continue
        cname = region_info["name"]

        for ct, chart_name in CHART_TYPES.items():
            if ct not in raw_data.get(country, {}):
                continue

            dates_data = raw_data[country][ct]
            # 按日期顺序整理为 [{rank: name}, ...]，共 8 个
            daily_data: List[Dict[int, str]] = []
            for d in DISPLAY_DATES:
                raw_map = dates_data.get(d, {})
                daily_data.append({int(k): v for k, v in raw_map.items()})

            # ── 创建 Sheet ────────────────────────────────────────────────
            sheet_name = f"{cname}_{chart_name}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_view.showGridLines = False

            # ── 行1：大标题 ───────────────────────────────────────────────
            title_text = (
                f"{cname} App Store 游戏{chart_name}"
                f" TOP{TOP}（3月16日-3月23日）"
            )
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
            tc = ws.cell(row=1, column=1, value=title_text)
            _style(tc, C_TITLE_BG, C_TITLE_FG, bold=True, size=11)
            ws.row_dimensions[1].height = 30

            # ── 行2：列头 ─────────────────────────────────────────────────
            rc = ws.cell(row=2, column=1, value="排名")
            _style(rc, C_HEAD_BG, C_HEAD_FG, bold=True)
            for ci, label in enumerate(date_labels, 2):
                hc = ws.cell(row=2, column=ci, value=label)
                _style(hc, C_HEAD_BG, C_HEAD_FG, bold=True)
            ws.row_dimensions[2].height = 26

            # ── 行3~(TOP+2)：数据 ─────────────────────────────────────────
            # 用第一天作为"基准"来判断是否新上榜
            first_day_names = set(daily_data[0].values()) if daily_data else set()
            last_day_dict   = daily_data[-1] if daily_data else {}

            for rank in range(1, TOP + 1):
                excel_row = rank + 2
                is_odd    = rank % 2 == 1
                row_bg    = C_ODD_BG if is_odd else C_EVEN_BG

                # 排名列
                rank_fg = (C_RANK1_FG if rank == 1
                           else C_RANK2_FG if rank == 2
                           else C_RANK3_FG if rank == 3
                           else C_RANK_FG)
                nc = ws.cell(row=excel_row, column=1, value=rank)
                _style(nc, C_RANK_BG, rank_fg, bold=(rank <= 3), size=11)
                ws.row_dimensions[excel_row].height = 22

                # 每天的游戏名列
                for di, day_dict in enumerate(daily_data):
                    orig_name    = day_dict.get(rank, "")
                    # 应用翻译（若有）
                    display_name = trans_map.get(orig_name, orig_name) if orig_name else ""
                    is_translated = bool(orig_name and display_name != orig_name)

                    # 最后一天：判断是否新上榜（首日未在榜 TOP 内）
                    is_last = (di == len(daily_data) - 1)
                    is_new  = is_last and orig_name and orig_name not in first_day_names

                    gc = ws.cell(row=excel_row, column=di + 2)

                    if is_new:
                        gc.value = f"★ {display_name}" if display_name else ""
                        _style(gc, C_NEW_BG, C_NEW_FG, bold=True)
                    elif is_translated:
                        gc.value = display_name
                        cell_fg = (C_RANK1_FG if rank == 1
                                   else C_RANK2_FG if rank == 2
                                   else C_RANK3_FG if rank == 3
                                   else C_TRANS_FG)
                        _style(gc, row_bg, cell_fg, bold=(rank <= 3))
                    else:
                        gc.value = display_name
                        cell_fg = (C_RANK1_FG if rank == 1
                                   else C_RANK2_FG if rank == 2
                                   else C_RANK3_FG if rank == 3
                                   else C_TEXT)
                        _style(gc, row_bg, cell_fg, bold=(rank <= 3 and bool(display_name)))

            # ── 底部来源行 ────────────────────────────────────────────────
            src_row = TOP + 3
            ws.merge_cells(start_row=src_row, start_column=1,
                           end_row=src_row, end_column=n_cols)
            sc = ws.cell(row=src_row, column=1,
                         value="资料来源：七麦数据 (qimai.cn)  |  绿色=已翻译  ★=新上榜(3/16未在榜)")
            _style(sc, C_TITLE_BG, "445566", size=9, h_align="left")
            ws.row_dimensions[src_row].height = 18

            # ── 列宽 ──────────────────────────────────────────────────────
            ws.column_dimensions["A"].width = 6          # 排名列
            for ci in range(2, n_cols + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 20   # 游戏名列

            ws.freeze_panes = "B3"   # 固定排名列和日期行

    wb.save(EXCEL_PATH)
    logger.info("[done] Excel 已保存: %s", EXCEL_PATH)


# ═════════════════════════════════════════════════════════════════════════════
# 四、汇总输出
# ═════════════════════════════════════════════════════════════════════════════

def print_summary(raw_data: dict, trans_map: Dict[str, str]):
    print()
    print("=" * 65)
    print("  [七麦数据爬取摘要]   2026-03-16 ~ 03-23")
    print("=" * 65)
    total_records = 0
    for country, region_info in REGIONS.items():
        if country not in raw_data:
            continue
        cname = region_info["name"]
        row_parts = []
        for ct, chart_name in CHART_TYPES.items():
            if ct not in raw_data.get(country, {}):
                continue
            days_ok = sum(
                1 for d in DISPLAY_DATES
                if raw_data[country][ct].get(d)
            )
            records = sum(
                len(raw_data[country][ct].get(d, {}))
                for d in DISPLAY_DATES
            )
            total_records += records
            row_parts.append(f"{chart_name} {days_ok}/8天({records}条)")
        if row_parts:
            print(f"  [{country.upper()}] {cname:8s}  " + "  |  ".join(row_parts))
    print("-" * 65)
    print(f"  合计记录数: {total_records}")
    print(f"  翻译词典 : {len(trans_map)} 个游戏名")
    if trans_map:
        print("  翻译示例 :")
        for orig, zh in list(trans_map.items())[:8]:
            print(f"    {orig}  ->  {zh}")
    print("=" * 65)
    print(f"\n  输出目录 : {OUTPUT_DIR}")
    print(f"  Excel   : {EXCEL_PATH.name}")
    print(f"  原始数据: {RAW_JSON_PATH.name}")
    print(f"  翻译缓存: {TRANS_JSON_PATH.name}")
    print()


# ═════════════════════════════════════════════════════════════════════════════
# 入口
# ═════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  [七麦历史榜单爬取 & 翻译]  2026-03-16 ~ 03-23")
    print("=" * 60)
    print(f"  日期: {DISPLAY_DATES[0]} ~ {DISPLAY_DATES[-1]}  ({len(DISPLAY_DATES)}天)")
    print(f"  地区: {len(REGIONS)} 个   榜单: 免费榜 + 畅销榜   TOP {TOP}")
    print("=" * 60)
    print()

    # ── Step 1: 爬取七麦历史数据 ─────────────────────────────────────────
    print("─── Step 1/3: 爬取七麦历史榜单数据 ───────────────────────")
    raw_data = fetch_all_data()

    # ── Step 2: 翻译所有非中文游戏名 ────────────────────────────────────
    print("\n─── Step 2/3: 翻译非中文游戏名 ────────────────────────────")
    non_chinese = collect_non_chinese_names(raw_data)
    logger.info("发现 %d 个非中文游戏名需要翻译", len(non_chinese))
    trans_map = translate_names(non_chinese)

    # ── Step 3: 生成 Excel ───────────────────────────────────────────────
    print("\n─── Step 3/3: 生成 Excel 报表 ──────────────────────────────")
    generate_excel(raw_data, trans_map)

    # ── 打印摘要 ─────────────────────────────────────────────────────────
    print_summary(raw_data, trans_map)
    print(">>> 全部完成！")


if __name__ == "__main__":
    main()
