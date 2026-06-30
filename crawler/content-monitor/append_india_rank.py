"""
追加印度地区 3月9日-3月15日 免费榜+畅销榜到现有 Excel，
并将 Sheet 中非中英文游戏名翻译为中文。
用法：cd D:/project/content-monitor && python append_india_rank.py
"""

import sys
import os
import re
import json
import logging
import time
from datetime import date, timedelta

sys.path.insert(0, os.path.dirname(__file__))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

EXCEL_FILE  = "D:/project/content-monitor/qimai_rank_0309-0315.xlsx"
START_DATE  = date(2026, 3, 9)
END_DATE    = date(2026, 3, 15)
COUNTRY     = "in"
CNAME       = "印度"
CHART_TYPES = ["free", "grossing"]
TOP         = 10


# ── 配色（与原脚本一致）──
C_TITLE_BG = "0F1824"
C_TITLE_FG = "F5B820"
C_HEAD_BG  = "1A2840"
C_HEAD_FG  = "F5B820"
C_RANK_BG  = "0F1824"
C_RANK_FG  = "6A8090"
C_ODD_BG   = "141E2E"
C_EVEN_BG  = "0F1824"
C_RANK1_FG = "FFD700"
C_RANK2_FG = "C0C0C0"
C_RANK3_FG = "CD7F32"
C_TEXT     = "C8D4E0"
C_NEW_FG   = "A78BFA"
C_NEW_BG   = "1E1535"
C_BORDER   = "243450"


def make_style_fn(border):
    from openpyxl.styles import PatternFill, Font, Alignment
    def style(c, bg, fg, bold=False, size=10, h_align="center"):
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(color=fg, bold=bold, size=size)
        c.alignment = Alignment(horizontal=h_align, vertical="center")
        c.border = border
    return style


def is_non_zh_en(name: str) -> bool:
    """判断是否含有非中文、非英文/数字/符号的字符（即需要翻译）。"""
    for ch in name:
        cp = ord(ch)
        # 跳过 ASCII、中文、全角标点
        if cp < 0x80:
            continue
        if 0x4E00 <= cp <= 0x9FFF:
            continue
        if 0x3000 <= cp <= 0x303F:
            continue
        # 其余非 ASCII 非中文 → 需要翻译
        return True
    return False


def translate_names(names: list) -> dict:
    """调用 qwen-max 翻译游戏名列表，返回 {原名: 中文名}。"""
    from qwen_client import get_qwen_client
    client = get_qwen_client()
    trans_map = {}
    batch_size = 40

    for i in range(0, len(names), batch_size):
        batch = names[i:i + batch_size]
        prompt = (
            "请将以下游戏名称翻译为中文（优先使用官方中文名，无官方名则音译或意译）。"
            "直接返回 JSON 对象，格式：{\"原名\": \"中文名\", ...}，不要 markdown，不要其他内容。\n\n"
            + "\n".join(f"- {n}" for n in batch)
        )
        logger.info("翻译第 %d 批（%d 个）...", i // batch_size + 1, len(batch))
        try:
            completion = client.chat.completions.create(
                model="qwen-max",
                messages=[{"role": "user", "content": prompt}],
                extra_body={"enable_thinking": False},
                max_tokens=2000,
                timeout=60,
            )
            raw = (completion.choices[0].message.content or "").strip()
            fence = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', raw)
            if fence:
                raw = fence.group(1).strip()
            batch_map = json.loads(raw)
            trans_map.update(batch_map)
            logger.info("  -> 成功翻译 %d 个", len(batch_map))
        except Exception as e:
            logger.warning("  翻译批次失败: %s", e)
        time.sleep(1)

    return trans_map


def build_india_sheets(wb):
    """爬取印度数据并写入新 Sheet，返回异动列表和 trans_map。"""
    import openpyxl
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    from scrapers.qimai_rank import fetch_rank_by_date, CHART_TYPES as CT_MAP

    thin   = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    style  = make_style_fn(border)

    days = []
    d = START_DATE
    while d <= END_DATE:
        days.append(d.strftime("%Y-%m-%d"))
        d += timedelta(days=1)

    date_range_label = f"{START_DATE.strftime('%m月%d日')}-{END_DATE.strftime('%m月%d日')}"

    all_anomalies = []
    # 收集所有游戏名（用于翻译检测）
    all_game_names = set()

    for chart_type in CHART_TYPES:
        chart_name = CT_MAP.get(chart_type, chart_type)
        sheet_name = f"{CNAME}_{chart_name}"

        # 如果 Sheet 已存在，先删除
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        logger.info("爬取 %s %s ...", CNAME, chart_name)

        # 抓取7天数据
        daily_data = []
        for day_str in days:
            logger.info("  %s", day_str)
            raw = fetch_rank_by_date(COUNTRY, chart_type, day_str, TOP)
            rank_to_name = {int(v): k for k, v in raw.items() if v is not None}
            daily_data.append(rank_to_name)
            all_game_names.update(rank_to_name.values())
            time.sleep(0.6)

        n_cols = 1 + len(days)

        # 标题行
        title = f"{CNAME} App Store 游戏{chart_name} TOP{TOP}（{date_range_label}）"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tc = ws.cell(row=1, column=1, value=title)
        style(tc, C_TITLE_BG, C_TITLE_FG, bold=True, size=11)
        ws.row_dimensions[1].height = 30

        # 列头
        rc = ws.cell(row=2, column=1, value="排名")
        style(rc, C_HEAD_BG, C_HEAD_FG, bold=True)
        for ci, day_str in enumerate(days, 2):
            m, day = day_str[5:7], day_str[8:]
            hc = ws.cell(row=2, column=ci, value=f"{int(m)}月{int(day)}日")
            style(hc, C_HEAD_BG, C_HEAD_FG, bold=True)
        ws.row_dimensions[2].height = 26

        # 数据行
        for rank in range(1, TOP + 1):
            row    = rank + 2
            is_odd = rank % 2 == 1
            row_bg = C_ODD_BG if is_odd else C_EVEN_BG
            rank_fg = (C_RANK1_FG if rank == 1 else
                       C_RANK2_FG if rank == 2 else
                       C_RANK3_FG if rank == 3 else C_RANK_FG)
            nc = ws.cell(row=row, column=1, value=rank)
            style(nc, C_RANK_BG, rank_fg, bold=(rank <= 3), size=11)
            ws.row_dimensions[row].height = 22

            for di, day_dict in enumerate(daily_data):
                game    = day_dict.get(rank, "")
                is_last = (di == len(daily_data) - 1)
                is_new  = is_last and game and (game not in daily_data[0].values())

                gc = ws.cell(row=row, column=di + 2, value=game)
                if is_new:
                    style(gc, C_NEW_BG, C_NEW_FG, bold=True)
                else:
                    cell_fg = (C_RANK1_FG if rank == 1 else
                               C_RANK2_FG if rank == 2 else
                               C_RANK3_FG if rank == 3 else C_TEXT)
                    style(gc, row_bg, cell_fg, bold=(rank <= 3 and bool(game)))

                # 异动检测
                if is_last and game:
                    if is_new:
                        all_anomalies.append({
                            "app": game, "region": CNAME, "chart": chart_name,
                            "change": "★新上榜", "rank": rank,
                        })
                    else:
                        first_rank = next(
                            (r for r, n in daily_data[0].items() if n == game), None
                        )
                        if first_rank and first_rank - rank >= 3:
                            all_anomalies.append({
                                "app": game, "region": CNAME, "chart": chart_name,
                                "change": f"↑{first_rank - rank}", "rank": rank,
                            })

        # 底部来源
        src_row = TOP + 3
        ws.merge_cells(start_row=src_row, start_column=1,
                       end_row=src_row, end_column=n_cols)
        sc = ws.cell(row=src_row, column=1, value="资料来源：七麦数据")
        style(sc, C_TITLE_BG, "445566", size=9, h_align="left")
        ws.row_dimensions[src_row].height = 18

        ws.column_dimensions["A"].width = 6
        for ci in range(2, n_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 16
        ws.freeze_panes = "B3"

    return all_anomalies, all_game_names


def translate_india_sheets(wb, trans_map: dict):
    """将印度 Sheet 中已翻译的名称回写。"""
    updated = 0
    for chart_name in ["免费榜", "畅销榜"]:
        sheet_name = f"{CNAME}_{chart_name}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                val = str(cell.value or "").strip()
                if val in trans_map and trans_map[val] != val:
                    cell.value = trans_map[val]
                    updated += 1
    return updated


def analyze_anomalies(anomalies: list):
    from qwen_client import analyze_rank_anomaly
    print(f"\n共检测到 {len(anomalies)} 条异动，AI 分析前 10 条...\n")
    for i, a in enumerate(anomalies[:10], 1):
        app    = a["app"]
        change = a["change"]
        rank   = a["rank"]
        chart  = a["chart"]
        print(f"[{i}] 分析 {app!r}  {chart} {change} -> 第{rank}名 ...")
        reason = analyze_rank_anomaly(
            app_name=app, region=CNAME, chart=chart,
            change=change, rank=rank,
        )
        a["reason"] = reason
        time.sleep(1)

    print("\n" + "=" * 65)
    print("印度榜单异动 AI 分析")
    print("=" * 65)
    for a in anomalies[:10]:
        print(f"\n【{a['app']}】{a['chart']}  {a['change']} -> 第{a['rank']}名")
        print(f"  {a.get('reason', '')}")
    print("=" * 65)


def main():
    import openpyxl

    logger.info("加载 Excel: %s", EXCEL_FILE)
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # 1. 爬取并写入印度 Sheet
    anomalies, all_game_names = build_india_sheets(wb)

    # 2. 检测需要翻译的游戏名（非中英文）
    to_translate = sorted({n for n in all_game_names if n and is_non_zh_en(n)})
    logger.info("需要翻译的游戏名：%d 个", len(to_translate))

    trans_map = {}
    if to_translate:
        trans_map = translate_names(to_translate)
        # 同步翻译异动列表中的游戏名
        for a in anomalies:
            a["app"] = trans_map.get(a["app"], a["app"])

    # 3. 回写翻译结果
    if trans_map:
        updated = translate_india_sheets(wb, trans_map)
        logger.info("翻译回写：更新 %d 个单元格", updated)

    # 4. 保存
    wb.save(EXCEL_FILE)
    logger.info("Excel 已保存: %s", EXCEL_FILE)

    # 5. 打印翻译对照
    if trans_map:
        print("\n--- 翻译对照 ---")
        for orig, zh in sorted(trans_map.items()):
            if orig != zh:
                try:
                    print(f"  {orig}  ->  {zh}")
                except UnicodeEncodeError:
                    pass

    # 6. AI 异动分析
    analyze_anomalies(anomalies)

    print(f"\n完成！文件：{os.path.abspath(EXCEL_FILE)}")


if __name__ == "__main__":
    main()
